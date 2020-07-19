# TODO Walk directory path added need to test, update all print to console and check logfile entries, testwith strongpassword Regex. 


'''Added memory tuneing to speed up time to completion of script(cut sript time to one quarter).  Added colored 'FOUND' flags to terminal output. Tested with two large spreadsheets.  Need to 
to test with a few .txt files.  Look at progress bars!(Now added need to look at some description text) '''

from datetime import datetime
from pathlib import Path
import os
import re
import chardet
import openpyxl
from openpyxl.utils import get_column_letter
from termcolor import colored
from tqdm import tqdm
import colorama


def timestamp():
    dateTimeObj = datetime.now()
    results_log = open('resultsLog.txt', 'a')
    results_log.write(str(dateTimeObj.strftime('%d/%m/%Y---%H:%M:%S   ')))
    results_log.close()


results_log = open('resultsLog.txt', 'a')
timestamp()
results_log.write('####Start txtFileRegexSearch Script####   \n')
results_log.close()

# TODO: open all *.txt files in a directory 
userRegex = re.compile(r'IB\w{10}')

# Iterate through folder structure under this path.
for folderName, subfolders, filenames in os.walk('C:\\test'):
    print(f'Searching Folder: {folderName} for .txt and .xlsx files')
    timestamp()
    results_log = open('resultsLog.txt', 'a')
    results_log.write(f'Searching Folder:  {folderName} for .txt and .xlsx files  \n' )
    results_log.close()

    p = Path(folderName)

    myList = list(p.glob('*.txt'))
    myExcel_FileList = list(p.glob('*.xlsx'))

    if not myList:
        print(f'No .txt files found in {str(p)}')
        timestamp()
        results_log = open('resultsLog.txt', 'a')
        results_log.write(f'No .txt files found in {str(p)}  \n')
        results_log.close()
    
    if not myExcel_FileList:
        print(f'No .xlsx files found in {str(p)}')
        timestamp()
        results_log = open('resultsLog.txt', 'a')
        results_log.write(f'No .xlsx files found in {str(p)}  \n')
        results_log.close()

    for testFileObject in p.glob('*.txt'):
            
        print('Txt file found: ' + str(testFileObject))
        timestamp()
        results_log = open('resultsLog.txt', 'a')
        results_log.write(f'Txt file found: {str(testFileObject)}  \n' )
        results_log.close()


    # TODO: Search through each line in  txt file for a user supplied regex expression


        rawdata = open(testFileObject, mode='rb').read()
        result = chardet.detect(rawdata)
        charenc = result['encoding']

        file_doc = open(testFileObject, encoding=charenc)
        file_lines = file_doc.readlines()
        file_hasReg = False

        for c, i in enumerate(file_lines):
            # print(userRegex.findall(i))
            if userRegex.findall(i):
                file_hasReg = True
                reg_Match = userRegex.findall(i)
                
                for regex_lineCount, txtLine_regMatch in enumerate(reg_Match, start=1):
                    print(colored('FOUND REGEX PATTERN ', 'green'), end='')
                    print(f'File {testFileObject}, Line: {c+1}, contains string: {regex_lineCount}--{txtLine_regMatch} ')
                    timestamp()
                    results_log = open('resultsLog.txt', 'a')
                    results_log.write(f'File {testFileObject}, Line: {c+1}, contains string: {regex_lineCount}--{txtLine_regMatch}    \n')
                    results_log.close()
               
                results_log.close()
            
        if file_hasReg == False:
            print('No Regex found in ' + str(testFileObject))
            timestamp()
            results_log = open('resultsLog.txt', 'a')
            results_log.write(f'No Regex found in  {str(testFileObject)}  \n' )
            results_log.close()

    for testExcelObject in p.glob('*.xlsx'):

        file_hasReg = False
        print(f'xlsx file found: {testExcelObject}')
        timestamp()
        results_log = open('resultsLog.txt', 'a')
        results_log.write(f'xlsx file found: {str(testExcelObject)}  \n')
        results_log.close()
        wb = openpyxl.load_workbook(filename= testExcelObject, read_only=True)
        sheet_Names = wb.sheetnames

        for unique_Sheet in sheet_Names:
            print(f'Searching SheetName {unique_Sheet}: in {testExcelObject}')
            timestamp()
            results_log = open('resultsLog.txt', 'a')
            results_log.write(f'Searching SheetName =: {unique_Sheet}: in {testExcelObject} \n')
            results_log.close()
            sheet_Search = wb[unique_Sheet]
                
            # Pre memory tuneing, new block below this one. 
            # for uni_row in range(1, sheet_Search.max_row+1):
            #     #myList.append([])  # Load new nested list index is referenced below.
            #     for cellVal2 in range(1, sheet_Search.max_column+1):
            #         print(uni_row)#(sheet_Search.cell(row=uni_row, column=cellVal2).value)
            #         search_Pattern = sheet_Search.cell(row=uni_row, column=cellVal2).value
            #         if search_Pattern:

            #             if userRegex.findall(search_Pattern):
            #                 reg_Match = userRegex.findall(search_Pattern)
            #                 column_letter = get_column_letter(cellVal2)
            #                 file_hasReg = True

            #                 for match_inList in reg_Match:
            #                     timestamp()
            #                     print(f'File {testExcelObject}, Sheet:{unique_Sheet}, row: {uni_row}, column: {column_letter} contains string: {match_inList}')
            #                     results_log = open('resultsLog.txt', 'a')
            #                     results_log.write(f'File {testExcelObject}, Sheet:{unique_Sheet}, row: {uni_row}, column: {column_letter} contains string: {match_inList} \n')
            #                     results_log.close()
            howmany_rows = wb[unique_Sheet].max_row
            for uni_row in tqdm(wb[unique_Sheet].rows, total=howmany_rows):
                for cellVal2 in uni_row:
                    search_Pattern = cellVal2.value
                    # if search_Pattern != None:
                    #     print(search_Pattern)
                    #     print(cellVal2.coordinate)
                    if search_Pattern:

                        if userRegex.findall(search_Pattern):
                            reg_Match = userRegex.findall(search_Pattern)
                            column_letter = cellVal2.coordinate
                            file_hasReg = True

                            for match_inList in reg_Match:
                                timestamp()
                                print(colored('\nFOUND REGEX PATTERN ', 'green'), end='')
                                print(f'File {testExcelObject}, Sheet:{unique_Sheet}, Cell: {column_letter} contains string: {match_inList}')
                                results_log = open('resultsLog.txt', 'a')
                                results_log.write(f'File {testExcelObject}, Sheet:{unique_Sheet}, Cell: {column_letter} contains string: {match_inList} \n')
                                results_log.close()




        if file_hasReg == False:
            print(f'No regex found in: {testExcelObject}')
            timestamp()
            results_log = open('resultsLog.txt', 'a')
            results_log.write(f'No regex found in: {str(testExcelObject)}  \n')
            results_log.close()


results_log = open('resultsLog.txt', 'a')
timestamp()
results_log.write('####End txtFileRegexSearch Script####  \n')
results_log.close()

