my_file = open('movies.txt', 'a')
my_file.write('Highlander, there can be only one\n')
my_file.write('Jaws, derda\n')
my_file.close()

my_file = open('movies.txt', 'r')
data = my_file.read()
my_file.close()
print(data)

for line in open('movies.txt'):
    print(line, end="")

import openpyxl

wb = openpyxl.load_workbook('example.xlsx')

sheet = wb['Sheet1']
print(sheet['A1'])

print(sheet['A1'].value)

c = sheet['B1']
print(c.value)

print('Row %s is %s' % (c.coordinate, c.value))

print(sheet['C1'].value)

print(sheet['A1'].value)

#--------------------------------------------------------

print(sheet.cell(row=1, column=2).value)

for i in range(1,8):
    print(i, sheet.cell(row=i, column=2).value)

#---------------------------------------------------------

print(sheet.max_row) 

print(sheet.max_column)

#-----------------------------------------------------

from openpyxl.utils import get_column_letter, column_index_from_string

print(get_column_letter(1))
print(get_column_letter(2))
print(get_column_letter(27))
print(get_column_letter(900))

print(get_column_letter(sheet.max_column))

print(column_index_from_string('A'))

print(column_index_from_string('AA'))

# #----------------------------------------------------------------------------
# Getting Rows and Columns from the SheetsYou can slice Worksheet objects to get all the Cell objects in a row, column, or rectangular area of the spreadsheet. Then you can loop over all the cells in the slice. Enter the following into the interactive shell:


wb = openpyxl.load_workbook('example.xlsx')

print(wb)

sheet = wb['Sheet1']

print(list(sheet['A1':'C3']))

for rowOfCellObjects in sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('------End Of Row------')

#---------------------------------------------------------------------

#-------------------------------------------------------------------------------

wb = openpyxl.load_workbook('example.xlsx')

sheet = wb.active

print(list(sheet.rows)[2])

# Gives a list of 7 tuples which contains all the rows, the above goes to row 3.
print(list(sheet.rows))

print(sheet.rows)

for cellObj in list(sheet.rows)[2]:
    print(cellObj.value)

# Using the rows attribute on a Worksheet object will give you a tuple of tuples. Each of these inner tuples represents a row, and contains the Cell objects in that row. The columns attribute also gives you a tuple of tuples, with each of the inner tuples containing the Cell objects in a particular column. For example.xlsx, since there are 7 rows and 3 columns, rows gives us a tuple of 7 tuples(each containing 3 Cell objects), and columns gives us a tuple of 3 tuples(each containing 7 Cell objects).To access one particular tuple, you can refer to it by its index in the larger tuple. For example, to get the tuple that represents column B, you use list(sheet.columns)[1]. To get the tuple containing the Cell objects in column A, you’d use list(sheet.columns)[0]. Once you have a tuple representing one row or column, you can loop through its Cell objects and print their values.
