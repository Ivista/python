from pathlib import Path
import openpyxl

wb = openpyxl.load_workbook('WebAppLog1.xlsx')
sheet = wb['Sheet1']

objFont1 = openpyxl.styles.Font(name='Calibri', bold=True, size=18, underline='single', color='0000FF')
objFont2 = openpyxl.styles.Font(name='Calibri', bold=True, size=18, underline='single', color='FF0000')

sheet.cell(row=sheet.max_row +1, column=1).value = 'Start of scan'
sheet.cell(row=sheet.max_row, column=1).font = objFont2


with open('C:\\ProgramData\\Scalable Software\\Scalable Agent\\Log\\ScalableAgent.log', 'r', encoding='utf-16-le') as scallog:
        data = scallog.readlines()

for c, i in enumerate(data):
        if 'Found a matching WebApp' in i:
                print(c+1)
                print(data[c-1])
                print(i)
                sheet.cell(row=sheet.max_row +1, column=1).value = data[c-1]
                sheet.cell(row=sheet.max_row +1, column=1).value = i



sheet.cell(row=sheet.max_row +1, column=1).value = 'End of scan'
sheet.cell(row=sheet.max_row, column=1).font = objFont1

sheet.column_dimensions['A'].width =210

wb.save('WebAppLog1.xlsx')