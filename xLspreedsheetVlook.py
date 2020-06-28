import openpyxl

wb = openpyxl.load_workbook('packages.xlsx')

print('hello')

sheet = wb['Sheet1']

#PACKAGE_UPDATES = {'GRP0122E': 'Demised', 'GRP046D3': 'Demised', 'GRP04A69': 'Demised'}

sheetone = wb['Sheet2']

sheetoneList = []

for col in range(1, sheetone.max_row +1):
    sheetoneList.append(sheetone.cell(row=col, column=1).value)

for rowNum in range(2, sheet.max_row +1):
    packageName = sheet.cell(row=rowNum, column=1).value
    if packageName in sheetoneList and sheet.cell(row=rowNum, column=5).value == 'Dave' :
        sheet.cell(row=rowNum, column=6).value = 'Demised'

#for dis in range(1, sheet.max_col):
objFont = openpyxl.styles.Font(name='Calibri', bold=True, size=18, italic=True)

for eader in range(1, sheet.max_column):
        sheet.cell(row=1, column=eader).font = objFont

sheet.column_dimensions['B'].width =40

wb.save('updatedpackages.xlsx')