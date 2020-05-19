from openpyxl.styles import Font
my_file = open('movies.txt', 'a')
my_file.write('Highlander, there can be only one\n')
my_file.write('Jaws, derda\n')
my_file.close()

my_file = open('movies.txt', 'r')
data = my_file.read()
my_file.close()
print(data)

for line in open('movies.txt'):
    print(line, end="")

import openpyxl

wb = openpyxl.load_workbook('example.xlsx')

sheet = wb['Sheet1']
print(sheet['A1'])

print(sheet['A1'].value)

c = sheet['B1']
print(c.value)

print('Row %s is %s' % (c.coordinate, c.value))

print(sheet['C1'].value)

print(sheet['A1'].value)

#--------------------------------------------------------

print(sheet.cell(row=1, column=2).value)

for i in range(1,8):
    print(i, sheet.cell(row=i, column=2).value)

#---------------------------------------------------------

print(sheet.max_row) 

print(sheet.max_column)

#-----------------------------------------------------

from openpyxl.utils import get_column_letter, column_index_from_string

print(get_column_letter(1))
print(get_column_letter(2))
print(get_column_letter(27))
print(get_column_letter(900))

print(get_column_letter(sheet.max_column))

print(column_index_from_string('A'))

print(column_index_from_string('AA'))

# #----------------------------------------------------------------------------
# Getting Rows and Columns from the SheetsYou can slice Worksheet objects to get all the Cell objects in a row, column, or rectangular area of the spreadsheet. Then you can loop over all the cells in the slice. Enter the following into the interactive shell:


wb = openpyxl.load_workbook('example.xlsx')

print(wb)

sheet = wb['Sheet1']

print(list(sheet['A1':'C3']))

for rowOfCellObjects in sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('------End Of Row------')

#---------------------------------------------------------------------

#-------------------------------------------------------------------------------

wb = openpyxl.load_workbook('example.xlsx')

sheet = wb.active

print(list(sheet.rows)[2])

# Gives a list of 7 tuples which contains all the rows, the above goes to row 3.
print(list(sheet.rows))

print(sheet.rows)

for cellObj in list(sheet.rows)[2]:
    print(cellObj.value)

# Using the rows attribute on a Worksheet object will give you a tuple of tuples. Each of these inner tuples represents a row, and contains the Cell objects in that row. The columns attribute also gives you a tuple of tuples, with each of the inner tuples containing the Cell objects in a particular column. For example.xlsx, since there are 7 rows and 3 columns, rows gives us a tuple of 7 tuples(each containing 3 Cell objects), and columns gives us a tuple of 3 tuples(each containing 7 Cell objects).To access one particular tuple, you can refer to it by its index in the larger tuple. For example, to get the tuple that represents column B, you use list(sheet.columns)[1]. To get the tuple containing the Cell objects in column A, you’d use list(sheet.columns)[0]. Once you have a tuple representing one row or column, you can loop through its Cell objects and print their values.


#Also check out updateproduce.py as this has a good row iter read and dict update


########-------------------------


wb = openpyxl.Workbook()
sheet = wb['Sheet']

italicFont24 = Font(size=24, italic=True, bold=True)

for colhead in range(1, 8):

    sheet.cell(row=1, column=colhead).font = italicFont24
    sheet.cell(row=1, column=colhead).value = 'Hello World!!!'
    #sheet.column_dimensions['A'].auto_size = True

    #sheet['A' + str(colhead)].font = italicFont24
    #sheet['A' + str(colhead)] = 'Hello World!'
sheet.column_dimensions['A'].width =
wb.save('Styles.xlsx')


#-----------------------------------------------------


wb = openpyxl.Workbook()

sheet = wb['Sheet']

fontObj1 = Font(color='00CCFFFF', name='Time New Roman', bold=True)
sheet['A1'].font = fontObj1
sheet['A1'] = 'Bold Times New Roman'

fontObj2 = Font(size=24, italic=True)
sheet['B3'].font = fontObj2
sheet['B3'] = '24 Point italic'

wb.save('Styles.xlsx')

# Here, we store a Font object in fontObj1 and then set the A1 Cell object’s font attribute to fontObj1. We repeat the process with another Font object to set the font of a second cell. After you run this code, the styles of the A1 and B3 cells in the spreadsheet will be set to custom font styles, as shown

#=---------Formulas


wb = openpyxl.Workbook()

sheet = wb.active

sheet['A1'] = 200
sheet['A2'] = 300
sheet['A3'] = '=SUM(A1:A2)'

wb.save('wrtieFormula.xlsx')

# The cells in A1 and A2 are set to 200 and 300, respectively. The value in cell A3 is set to a formula that sums the values in A1 and A2. When the spreadsheet is opened in Excel, A3 will display its value as 500.Excel formulas offer a level of programmability for spreadsheets but can quickly become unmanageable for complicated tasks. For example, even if you’re deeply familiar with Excel formulas, it’s a headache to try to decipher what = IFERROR(TRIM(IF(LEN(VLOOKUP(F7, Sheet2!$A$1: $B$10000, 2, FALSE)) > 0, SUBSTITUTE(VLOOKUP(F7, Sheet2!$A$1: $B$10000, 2, FALSE), " ", ""), "")), "") actually does. Python code is much more readable.


#----------------------------------------------------------------------------------


#------------------------------------------------------
#ADJUSTING ROWS AND COLUMNS

# In Excel, adjusting the sizes of rows and columns is as easy as clicking and dragging the edges of a row or column header. But if you need to set a row or column’s size based on its cells’ contents or if you want to set sizes in a large number of spreadsheet files, it will be much quicker to write a Python program to do it.

# Rows and columns can also be hidden entirely from view. Or they can be “frozen” in place so that they are always visible on the screen and appear on every page when the spreadsheet is printed(which is handy for headers).

# Setting Row Height and Column Width
# Worksheet objects have row_dimensions and column_dimensions attributes that control row heights and column widths. Enter this into the interactive shell:


wb = openpyxl.Workbook()

sheet = wb.active

sheet['A1'] = 'Tall Row'
sheet['B2'] = 'Wide column'
#set the hieght and width

sheet.row_dimensions[1].hieght = 70
sheet.column_dimensions['B'].width = 20
wb.save('dimensions.xlsx')

# Merging and Unmerging CellsA rectangular area of cells can be merged into a single cell with the merge_cells() sheet method. Enter the following into the interactive shell:


wb = openpyxl.Workbook()

sheet = wb.active

sheet.merge_cells('A1:D3')

sheet['A1'] = '12 Cells merged together'
sheet.merge_cells('C5:D5')
sheet['C5'] = 'Two merged cells'

wb.save('merged.xlsx')


#See below to unmerge cells =------------------------------


wb = openpyxl.load_workbook('merged.xlsx')

sheet = wb.active

sheet.unmerge_cells('A1:D3')
sheet.unmerge_cells('C5:D5')
wb.save('merged.xlsx')

#-------------------------------------------------------------
#Freeze panes


wb = openpyxl.load_workbook('produceSales.xlsx')

sheet = wb.active

sheet.freeze_panes = 'A2'

wb.save('freezeExample.xlsx')

#-------------------Bar Chart-----------------------------------


wb = openpyxl.Workbook()

sheet = wb.active

for i in range(1, 11):
    sheet['A' + str(i)] = i

refObj = openpyxl.chart.Reference(
    sheet, min_col=1, min_row=1, max_col=1, max_row=10)

seriesObj = openpyxl.chart.Series(refObj, title='My new series')

chartobj = openpyxl.chart.BarChart(overlap=True)

chartobj.title = 'My Chart'

chartobj.append(seriesObj)

sheet.add_chart(chartobj, 'C5')
wb.save('sampleChart.xlsx')


#----------------------------------------------------
# Multiplication Table Maker
# Create a program multiplicationTable.py that takes a number N from the command line and creates an N×N multiplication table in an Excel spreadsheet. For example, when the program is run like this:

wb = openpyxl.Workbook()

sheet = wb.active

for colandrow in range(2, 8):
    sheet.cell(row=1, column=colandrow).value = colandrow - 1
    sheet.cell(row=colandrow, column=1).value = colandrow - 1


max_rows = sheet.max_row + 1
max_cells = sheet.max_column

for cellVal in range(2, max_rows):
    for cellVal2 in range(2, max_rows):

        val1 = sheet.cell(row=cellVal, column=1).value
        val2 = sheet.cell(row=1, column=cellVal2).value
        val3 = val1 * val2
        sheet.cell(row=cellVal, column=cellVal2).value = val3


print('hello')

wb.save('workingproj.xlsx')
#------------------------------------------------------------------------------
# Create a program blankRowInserter.py that takes two integers and a filename string as command line arguments. Let’s call the first integer N and the second integer M. Starting at row N, the program should insert M blank rows into the spreadsheet. For example, when the program is run like this:


wbinput = openpyxl.load_workbook('spreadsheet-A.xlsx')
sheetInput = wbinput.active

wboutput = openpyxl.Workbook()
sheetOutput = wboutput.active

n = 3  # Startng at row n
m = 2  # This many rows shoud be inserted.

tot = n+m

for row1 in range(1, n):
    for row2 in range(1, sheetInput.max_column):
        sheetOutput.cell(row=row1, column=row2).value = sheetInput.cell(
            row=row1, column=row2).value

for row3 in range(tot, sheetInput.max_row): #row 3 is the anchor for the row and should be used in the nested for loop below.
    for row4 in range(tot, sheetInput.max_column+tot):
        sheetOutput.cell(row=row3, column=row4-tot +
                         1).value = sheetInput.cell(row=row3-m, column=row4-tot+1).value


wboutput.save('outPutforRowInsert.xlsx')


#-----------------------------------------------
#loads one sheets rows into a list then switches them around and writes to a new workbook


wb = openpyxl.load_workbook('spreadsheet-A.xlsx')
wbNew = openpyxl.Workbook()

sheet = wb.active
sheetNew = wbNew.active

#myList =[['ITEM', 'SOLD'], ['EggPlant', 334], ['Cucumber', 252]]
#print(myList[0][1])

myList = []
myDict = {}

for cellVal in range(1, sheet.max_row+1):
    myList.append([])  # Load new nested list index is referenced below.
    for cellVal2 in range(1, sheet.max_column+1):
        print(sheet.cell(row=cellVal, column=cellVal2).value)
        myList[cellVal-1].append(sheet.cell(row=cellVal,
                                            column=cellVal2).value)  # new nested list loaded with data

for c, sub in enumerate(myList):
    print(c)
    for d, measure in enumerate(sub):
        print(measure)
        sheetNew.cell(column=c+1, row=d+1).value = measure
        if measure == 'ALUMNI':  # If value of cess is this load dict
            colval = c+1
            rowval = d+1
            # Load dict with key value pairs col = key, row = value
            myDict[colval] = rowval

for key, value in myDict.items():
    # If column in dict print match below last row value
    sheetNew.cell(column=key, row=sheetNew.max_row+1).value = 'MATCH'

wbNew.save('testoutput.xlsx')
